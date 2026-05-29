#!/usr/bin/env python3
"""
build_kronologia.py — Kronologia interaktiboaren eraikitzailea.

100 data-klabeen Excel-etik (content/data/) abiatuta, denbora-lerro
interaktiboa sortzen du: kronologia.html (datuak HTML barruan txertatuta,
beraz GitHub Pages-en zein lokalean dabil, fetch-ik gabe).

Erabilera:   python3 build_kronologia.py
Behar du:    pip install openpyxl
"""
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Falta openpyxl. Exekutatu: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).parent
XLSX = ROOT / "content" / "data" / "100_data_klabeak_1800_2020_Historia_BTX2.xlsx"
OUT_HTML = ROOT / "kronologia.html"
OUT_JOKOA = ROOT / "jokoa.html"
OUT_JSON = ROOT / "assets" / "data" / "kronologia.json"
CACHE_V = 9  # bumpea CSS/JS aldatzean

# Izenburuan urtea agertzen duten datak jokotik kanpo (adib. "1812ko Konstituzioa").
# Urtea euskal atzizkiari lotuta egon daiteke ("1837ko"), beraz \b ez da nahikoa.
YEAR_IN_TITLE = re.compile(r"(?<!\d)(1[89]\d{2}|20\d{2})(?!\d)")

# --- Denbora-lerroaren konfigurazioa --------------------------------------
AXIS_MIN = 1808
AXIS_MAX = 2021

# 7 garai jarrai (atzeko planoko bandak + bizkor-jauziak)
ERAS = [
    {"s": 1808, "e": 1833, "l": "Antzinako Erregimenaren krisia", "c": "#ef4444"},
    {"s": 1833, "e": 1874, "l": "Estatu liberala eta karlistadak", "c": "#f59e0b"},
    {"s": 1874, "e": 1931, "l": "Berrezarkuntza eta industrializazioa", "c": "#14b8a6"},
    {"s": 1931, "e": 1939, "l": "II. Errepublika eta Gerra Zibila", "c": "#b91c1c"},
    {"s": 1939, "e": 1975, "l": "Frankismoa", "c": "#64748b"},
    {"s": 1975, "e": 1982, "l": "Trantsizioa", "c": "#3b82f6"},
    {"s": 1982, "e": 2021, "l": "Demokrazia, Europa eta XXI. mendea", "c": "#0891b2"},
]

# Hiru eremuak (lerro/karrilen koloreak)
EREMU = {
    "Espainia":      {"key": "es", "name": "Espainia",      "main": "#6d28d9", "soft": "#ede9fe", "ink": "#5b21b6"},
    "Euskal Herria": {"key": "eh", "name": "Euskal Herria", "main": "#0e9f6e", "soft": "#d1fae5", "ink": "#065f46"},
    "Mundua":        {"key": "mu", "name": "Mundua",        "main": "#ea580c", "soft": "#ffedd5", "ink": "#9a3412"},
}

# 10 prozesu nagusiak (Excel-eko "Prozesu mapa" orritik; ikasketarako testuingurua)


def era_index(year):
    for i, era in enumerate(ERAS):
        if era["s"] <= year < era["e"]:
            return i
    return len(ERAS) - 1


def parse_years(raw):
    if raw is None:
        return []
    return [int(m) for m in re.findall(r"\b(1[89]\d{2}|20\d{2})\b", str(raw))]


def split_list(raw):
    if raw is None:
        return []
    parts = re.split(r"[;\n]", str(raw))
    return [p.strip() for p in parts if p and p.strip()]


def norm(s):
    """Bilaketarako: minuskula + azentuak kendu."""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower()


def load_records():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["100 data klabe"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])

    def col(row, name):
        return row[headers.index(name)]

    recs = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        year = int(col(row, "Urtea"))
        eremu_raw = (col(row, "Eremu nagusia") or "").strip()
        eremu = EREMU.get(eremu_raw, EREMU["Mundua"])
        title = (col(row, "Gertakaria") or "").strip()
        kontz = split_list(col(row, "Kontzeptu giltzak"))
        azalpena = (col(row, "Azalpen didaktikoa") or "").strip()
        aldi = (col(row, "Aldi historikoa") or "").strip()
        ardatz = (col(row, "Ardatz tematikoa") or "").strip()
        rec = {
            "id": int(col(row, "ID")),
            "y": year,
            "d": str(col(row, "Data") or year),
            "t": title,
            "e": eremu["key"],
            "ee": (col(row, "Eragin-eremua") or "").strip(),
            "al": aldi,
            "ax": ardatz,
            "ka": azalpena,
            "zk": (col(row, "Zergatik da klabea?") or "").strip(),
            "lo": parse_years(col(row, "Lotutako datak")),
            "p": "H" if (col(row, "PAU/EBaU balioa") or "").strip().startswith("Hand") else "E",
            "kg": kontz,
            "dm": (col(row, "Data mota") or "").strip(),
            "u": (col(row, "Iturri URL orientagarria") or "").strip(),
            "b": era_index(year),
        }
        # Bilaketa-indizea (azenturik gabe)
        rec["s"] = norm(" ".join([title, " ".join(kontz), azalpena, aldi, ardatz, str(year), rec["dm"]]))
        recs.append(rec)
    recs.sort(key=lambda r: (r["y"], r["id"]))
    return recs


def load_prozesuak():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Prozesu mapa"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        out.append({
            "p": str(r[0]).strip(),
            "tartea": str(r[1] or "").strip(),
            "datak": str(r[2] or "").strip(),
            "hari": str(r[3] or "").strip(),
            "erabilera": str(r[4] or "").strip(),
        })
    return out


def main():
    if not XLSX.exists():
        print(f"❌ Ez dago Excel-a: {XLSX}", file=sys.stderr)
        sys.exit(1)

    records = load_records()
    prozesuak = load_prozesuak()

    config = {
        "axisMin": AXIS_MIN,
        "axisMax": AXIS_MAX,
        "eras": ERAS,
        "eremu": {v["key"]: {"name": v["name"], "main": v["main"], "soft": v["soft"], "ink": v["ink"]}
                  for v in EREMU.values()},
        "records": records,
        "prozesuak": prozesuak,
    }

    # JSON ere gorde (gardentasuna / berrerabilpena)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(config, ensure_ascii=False, indent=1), encoding="utf-8")

    data_json = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
    n = len(records)
    counts = {"es": 0, "eh": 0, "mu": 0}
    for r in records:
        counts[r["e"]] += 1
    handia = sum(1 for r in records if r["p"] == "H")
    year = datetime.now().year

    html = PAGE.format(
        v=CACHE_V,
        data_json=data_json,
        n=n,
        es=counts["es"],
        eh=counts["eh"],
        mu=counts["mu"],
        handia=handia,
        current_year=year,
    )
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"✓ kronologia.html  ({n} data-klabe · {counts['es']} Espainia / {counts['eh']} EH / {counts['mu']} Mundua)")
    print(f"✓ assets/data/kronologia.json")

    # --- Jokoa: izenburuan urterik EZ duten datak (urtea ezin da agerian egon) ---
    pool = [{"id": r["id"], "t": r["t"], "y": r["y"], "e": r["e"]}
            for r in records if not YEAR_IN_TITLE.search(r["t"])]
    jokoa_cfg = {
        "pool": pool,
        "eremu": config["eremu"],
    }
    jokoa_json = json.dumps(jokoa_cfg, ensure_ascii=False, separators=(",", ":"))
    OUT_JOKOA.write_text(JOKOA.format(v=CACHE_V, data_json=jokoa_json, npool=len(pool), current_year=year), encoding="utf-8")
    print(f"✓ jokoa.html  ({len(pool)} data jokorako; {n - len(pool)} baztertuta izenburuan urtea dutelako)")


# --- HTML plantilla (str.format; {{ }} = literal giltza) -------------------
PAGE = r"""<!DOCTYPE html>
<html lang="eu">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Kronologia interaktiboa · 100 data-klabe (1808–2020) | Historia BTX2</title>
  <meta name="description" content="1808tik 2020ra arteko 100 data-klabe interaktiboak hiru eremutan (Espainia, Euskal Herria eta Mundua): iragazkiak, bilaketa, azalpen didaktikoak eta lotutako datak EBAU/PAU prestatzeko, euskaraz." />
  <link rel="canonical" href="https://btx2-eus.github.io/historia-btx2/kronologia.html" />
  <meta property="og:locale" content="eu_ES" />
  <meta property="og:type" content="website" />
  <meta property="og:site_name" content="Historia BTX2" />
  <meta property="og:title" content="Kronologia interaktiboa · 100 data-klabe (1808–2020)" />
  <meta property="og:description" content="1808–2020 arteko 100 data-klabe interaktiboak hiru eremutan: Espainia, Euskal Herria eta Mundua. EBAU/PAU prestatzeko, euskaraz." />
  <meta property="og:url" content="https://btx2-eus.github.io/historia-btx2/kronologia.html" />
  <meta name="twitter:card" content="summary" />
  <meta name="twitter:title" content="Kronologia interaktiboa · 100 data-klabe" />
  <meta name="twitter:description" content="1808–2020 arteko 100 data-klabe interaktiboak hiru eremutan. EBAU/PAU prestatzeko, euskaraz." />
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet" />
  <link rel="stylesheet" href="assets/css/styles.css?v={v}" />
  <link rel="stylesheet" href="assets/css/kronologia.css?v={v}" />
  <link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Crect width='100' height='100' rx='24' fill='%235b2be0'/%3E%3Ctext x='50' y='70' font-size='58' text-anchor='middle' fill='white' font-family='sans-serif' font-weight='bold'%3EH%3C/text%3E%3C/svg%3E" />
</head>
<body class="kx-body">
  <a class="skip-link" href="#main">Edukira saltatu</a>

  <header class="nav">
    <div class="container">
      <a class="brand" href="index.html"><span class="logo">H</span><span>Historia BTX2<small>Espainiako Historia · 2. Batxilergoa</small></span></a>
      <button class="nav-toggle" aria-label="Menua" aria-expanded="false"><span></span><span></span><span></span></button>
      <nav class="nav-links">
        <a href="index.html#denbora-lerroa">Denbora-lerroa</a>
        <a href="index.html#gaiak">Gaiak</a>
        <a href="index.html#tresnak">Tresnak</a>
        <a class="cta" href="index.html">← Hasiera</a>
      </nav>
    </div>
  </header>

  <section class="kx-hero" id="main" tabindex="-1">
    <div class="container">
      <span class="hero-eyebrow">★ Kronologia interaktiboa · EBAU/PAU</span>
      <h1>100 data-klabe, <span class="grad-text">hiru begiradatan</span></h1>
      <p class="lead">1808tik 2020ra arteko ehun data giltzarri, aldi berean hiru eremutan irakurrita: <b>Espainia</b>, <b>Euskal Herria</b> eta <b>Mundua</b>. Sakatu data batean azalpen didaktikoa, zergatia eta lotutako datak ikusteko.</p>
      <div class="kx-hero-stats">
        <div class="kx-stat"><strong>{n}</strong><span>data-klabe</span></div>
        <div class="kx-stat es"><strong>{es}</strong><span>Espainia</span></div>
        <div class="kx-stat eh"><strong>{eh}</strong><span>Euskal Herria</span></div>
        <div class="kx-stat mu"><strong>{mu}</strong><span>Mundua</span></div>
        <div class="kx-stat"><strong>{handia}</strong><span>klabe handi</span></div>
      </div>
    </div>
  </section>

  <main class="kx-main">
    <div class="container-wide">

      <!-- Kontrol-barra -->
      <div class="kx-controls" role="region" aria-label="Iragazkiak eta kontrolak">
        <div class="kx-search">
          <span class="kx-search-ic" aria-hidden="true">🔎</span>
          <input id="kx-q" type="search" placeholder="Bilatu gertaera, kontzeptua, urtea…" aria-label="Bilatu kronologian" autocomplete="off" />
          <button id="kx-q-clear" class="kx-q-clear" aria-label="Bilaketa garbitu" hidden>✕</button>
        </div>

        <div class="kx-filters">
          <div class="kx-chipset" role="group" aria-label="Eremuak">
            <button class="kx-chip es is-on" data-eremu="es" aria-pressed="true"><span class="kx-dot"></span>Espainia</button>
            <button class="kx-chip eh is-on" data-eremu="eh" aria-pressed="true"><span class="kx-dot"></span>Euskal Herria</button>
            <button class="kx-chip mu is-on" data-eremu="mu" aria-pressed="true"><span class="kx-dot"></span>Mundua</button>
          </div>
          <button class="kx-toggle" id="kx-pau" aria-pressed="false"><span class="kx-star">★</span> Klabe handiak soilik</button>
        </div>

        <div class="kx-tools">
          <span class="kx-count" id="kx-count" aria-live="polite"></span>
          <div class="kx-zoom" role="group" aria-label="Zooma">
            <button id="kx-zoom-out" aria-label="Urrundu (zoom out)">−</button>
            <button id="kx-zoom-fit" aria-label="Doitu pantailara">Doitu</button>
            <button id="kx-zoom-in" aria-label="Hurbildu (zoom in)">+</button>
          </div>
          <button class="kx-reset" id="kx-reset">↺ Berrezarri</button>
        </div>
      </div>

      <!-- Garai bizkor-jauziak -->
      <div class="kx-eranav" id="kx-eranav" role="group" aria-label="Garaira jauzi"></div>

      <!-- Denbora-lerroa -->
      <div class="kx-stage">
        <div class="kx-scroll" id="kx-scroll" tabindex="0" aria-label="Denbora-lerroa, alboetara lerratu daiteke">
          <div class="kx-canvas" id="kx-canvas">
            <div class="kx-eras" id="kx-eras" aria-hidden="true"></div>
            <div class="kx-eralabels" id="kx-eralabels" aria-hidden="true"></div>
            <div class="kx-ruler" id="kx-ruler" aria-hidden="true"></div>
            <div class="kx-lanes" id="kx-lanes"></div>
          </div>
        </div>
        <div class="kx-empty" id="kx-empty" hidden>Ez da emaitzarik aurkitu. Saiatu iragazkiak aldatzen edo bilaketa garbitzen.</div>
        <p class="kx-hint">💡 Lerratu alboetara · sakatu data bat xehetasunetarako · erabili ← → teklak nabigatzeko</p>
      </div>

      <!-- Prozesu nagusiak -->
      <section class="kx-prozesuak">
        <h2>Hari nagusiak, 1808–2020</h2>
        <p class="kx-prozesuak-lead">Ehun datak hamar prozesu historikotan antolatzen dira. Hauek dira EBAU/PAUrako kausa-kateak eta narratiba nagusiak.</p>
        <div class="kx-prozesuak-grid" id="kx-prozesuak"></div>
      </section>
    </div>
  </main>

  <!-- Xehetasun-panela -->
  <aside class="kx-drawer" id="kx-drawer" role="dialog" aria-modal="false" aria-labelledby="kx-d-title" hidden>
    <div class="kx-drawer-head">
      <div class="kx-d-year" id="kx-d-year"></div>
      <button class="kx-drawer-close" id="kx-drawer-close" aria-label="Itxi">✕</button>
    </div>
    <div class="kx-drawer-body" id="kx-drawer-body"></div>
  </aside>
  <div class="kx-scrim" id="kx-scrim" hidden></div>

  <footer class="footer">
    <div class="container">
      <a class="brand" href="index.html"><span class="logo">H</span><span>Historia BTX2<small>Luken San Sebastián Alkorta</small></span></a>
      <nav class="foot-links">
        <a href="index.html#denbora-lerroa">Denbora-lerroa</a>
        <a href="kronologia.html">Kronologia</a>
        <a href="index.html#gaiak">Gaiak</a>
        <a href="index.html#tresnak">Tresnak</a>
      </nav>
      <div class="foot-bottom">
        <span>Kronologia interaktiboa · 100 data-klabe · 1808–2020</span>
        <span>© <span data-year>{current_year}</span> · Luken San Sebastián Alkorta · <a href="https://creativecommons.org/licenses/by-sa/4.0/" target="_blank" rel="noopener license">CC BY-SA 4.0</a></span>
      </div>
    </div>
  </footer>

  <script id="kx-data" type="application/json">{data_json}</script>
  <script src="assets/js/app.js?v={v}"></script>
  <script src="assets/js/kronologia.js?v={v}"></script>
</body>
</html>
"""


JOKOA = r"""<!DOCTYPE html>
<html lang="eu">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Datak ordenatzeko jokoa · Historia BTX2</title>
  <meta name="description" content="Datak ordenatzeko joko interaktiboa: aukeratu 5, 7, 10 edo 15 gertaera, ordenatu zaharrenetik berrienera eta programak zuzenduko du. EBAU/PAU errepasorako, euskaraz." />
  <link rel="canonical" href="https://btx2-eus.github.io/historia-btx2/jokoa.html" />
  <meta property="og:locale" content="eu_ES" />
  <meta property="og:type" content="website" />
  <meta property="og:site_name" content="Historia BTX2" />
  <meta property="og:title" content="Datak ordenatzeko jokoa · Historia BTX2" />
  <meta property="og:description" content="Aukeratu 5, 7, 10 edo 15 gertaera eta ordenatu zaharrenetik berrienera. EBAU/PAU errepasorako." />
  <meta property="og:url" content="https://btx2-eus.github.io/historia-btx2/jokoa.html" />
  <meta name="twitter:card" content="summary" />
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet" />
  <link rel="stylesheet" href="assets/css/styles.css?v={v}" />
  <link rel="stylesheet" href="assets/css/jokoa.css?v={v}" />
  <link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Crect width='100' height='100' rx='24' fill='%235b2be0'/%3E%3Ctext x='50' y='70' font-size='58' text-anchor='middle' fill='white' font-family='sans-serif' font-weight='bold'%3EH%3C/text%3E%3C/svg%3E" />
</head>
<body class="jk-body">
  <a class="skip-link" href="#main">Edukira saltatu</a>

  <header class="nav">
    <div class="container">
      <a class="brand" href="index.html"><span class="logo">H</span><span>Historia BTX2<small>Espainiako Historia · 2. Batxilergoa</small></span></a>
      <button class="nav-toggle" aria-label="Menua" aria-expanded="false"><span></span><span></span><span></span></button>
      <nav class="nav-links">
        <a href="index.html#denbora-lerroa">Denbora-lerroa</a>
        <a href="kronologia.html">Kronologia</a>
        <a href="jokoa.html">Jokoa</a>
        <a class="cta" href="index.html">← Hasiera</a>
      </nav>
    </div>
  </header>

  <section class="jk-hero" id="main" tabindex="-1">
    <div class="container">
      <span class="hero-eyebrow">🎯 Jokoa · EBAU/PAU errepasoa</span>
      <h1>Ordenatu <span class="grad-text">datak</span></h1>
      <p class="lead">Gertaerak nahastuta agertuko zaizkizu, izena bakarrik. Ordenatu <b>zaharrenetik (goian) berrienera (behean)</b> eta sakatu «Zuzendu». Urtea ez da agertuko zuzendu arte.</p>
    </div>
  </section>

  <main class="jk-main">
    <div class="container">

      <div class="jk-setup" id="jk-setup">
        <div class="jk-setup-row">
          <span class="jk-setup-label">Zenbat data?</span>
          <div class="jk-count-chips" id="jk-count-chips" role="group" aria-label="Data kopurua">
            <button class="jk-count is-on" data-n="5" aria-pressed="true">5</button>
            <button class="jk-count" data-n="7" aria-pressed="false">7</button>
            <button class="jk-count" data-n="10" aria-pressed="false">10</button>
            <button class="jk-count" data-n="15" aria-pressed="false">15</button>
          </div>
        </div>
        <button class="btn btn-primary jk-new" id="jk-new">🎲 Joko berria</button>
      </div>

      <div class="jk-board" id="jk-board" hidden>
        <div class="jk-axis" aria-hidden="true">
          <span class="jk-axis-top">▲ Zaharrena</span>
          <span class="jk-axis-line"></span>
          <span class="jk-axis-bot">Berriena ▼</span>
        </div>
        <ol class="jk-list" id="jk-list"></ol>
      </div>

      <div class="jk-actions" id="jk-actions" hidden>
        <button class="btn btn-primary" id="jk-check">✓ Zuzendu</button>
        <button class="btn btn-soft" id="jk-shuffle">🔀 Berriz nahastu</button>
        <button class="btn btn-soft" id="jk-again" hidden>🎲 Beste joko bat</button>
      </div>

      <div class="jk-result" id="jk-result" hidden aria-live="polite"></div>

      <div class="jk-solution" id="jk-solution" hidden>
        <h2>Ordena zuzena</h2>
        <ol class="jk-sol-list" id="jk-sol-list"></ol>
      </div>

      <p class="jk-hint">💡 Arrastatu txartelak edo erabili ↑ ↓ botoiak ordenatzeko. {npool} gertaeren artean (izenburuan urtea dutenak kanpoan utzita).</p>
    </div>
  </main>

  <footer class="footer">
    <div class="container">
      <a class="brand" href="index.html"><span class="logo">H</span><span>Historia BTX2<small>Luken San Sebastián Alkorta</small></span></a>
      <nav class="foot-links">
        <a href="index.html#denbora-lerroa">Denbora-lerroa</a>
        <a href="kronologia.html">Kronologia</a>
        <a href="jokoa.html">Jokoa</a>
        <a href="index.html#gaiak">Gaiak</a>
      </nav>
      <div class="foot-bottom">
        <span>Datak ordenatzeko jokoa · Historia BTX2</span>
        <span>© <span data-year>{current_year}</span> · Luken San Sebastián Alkorta · <a href="https://creativecommons.org/licenses/by-sa/4.0/" target="_blank" rel="noopener license">CC BY-SA 4.0</a></span>
      </div>
    </div>
  </footer>

  <script id="jk-data" type="application/json">{data_json}</script>
  <script src="assets/js/app.js?v={v}"></script>
  <script src="assets/js/jokoa.js?v={v}"></script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
